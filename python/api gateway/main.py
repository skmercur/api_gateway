
from flask import Flask,request
import openpyxl
import pandas as pd
import requests
from prettytable import PrettyTable
from flask_cors import CORS
import os

app = Flask(__name__)
cors = CORS(app, resources={r"/api/*": {"origins": "*"}})
message = input("Would you like to view or edit route? (y/n) ")
x = PrettyTable(field_names=['index', 'service', 'endpoint'])
routings = pd.read_excel('routing.xlsx')
if message == "y":
    os.system("clear")
    for (index,row) in routings.iterrows():
        x.add_row([index,row['service'],row['endpoint']])
    print(x)
    message2 = input("Would you like to add a route? (y/n) ")
    if message2 == "y":
        os.system("clear")
       
        service = input("What service would you like to add? ")
        endpoint = input("What endpoint would you like to add? ")
        routings.loc[len(routings)] = [service,endpoint]
        routings.to_excel('routing.xlsx',index=False)
        print("Route added")
    edit = input("Would you like to edit the endpoint? (y/n) ")
        
        
    if edit == "y":
        index = input("select index of endpoint to edit exemple 1 ")
        wb = openpyxl.load_workbook('routing.xlsx')
        sheet = wb.active
        service  = input("What service would you like to add? ")
        sheet.cell(row=int(index)+2, column=1).value = service
        endpoint = input("What endpoint would you like to add? ")
        sheet.cell(row=int(index)+2, column=2).value = endpoint
        wb.save('routing.xlsx')
        

        
        

else:
    routings = pd.read_excel('routing.xlsx')
    @app.route('/')
    def mainSecreen():
        return "Hello World"
    @app.route("/api/v1/<string:service>",methods=['GET','POST'])
    def endpoints(service):
        if request.method == "POST":
            selected_routing = routings[routings['service'] == service]
            request_data = requests.post(selected_routing['endpoint'].iloc[0],json=request.json)
            return {"endpoint":service,"request_body":request.json,"response":request_data.json()}
        if request.method == "GET":

            request_data = requests.get(selected_routing['endpoint'].iloc[0],json=request.json)
            return {"endpoint":service,"request_body":request.json,"response":request_data.json()}
        else:
            return {"Message":"Method not allowed"}






if __name__ == '__main__':
    app.run( host='0.0.0.0',port=5000)